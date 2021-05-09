from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from rich.console import Console
from rich.table import Table
from selenium import webdriver
from datetime import datetime
from art import *
import openpyxl
import calendar
import pandas
import cursor
import json
import time
import os
import re

def fetchDataFromJSON(fileName):
	with open(fileName) as file:
		data = json.load(file)
	return data

def sendDataToJSON(fileName, data):
	with open(fileName, 'w') as file:
		json.dump(data, file, indent = 4)

data = fetchDataFromJSON('data.json')
profilePath = data['dir']['profilePath']

console = Console()
chromeOptions = Options()
chromeOptions.add_argument("--disable-extensions")
chromeOptions.add_argument("--disable-popup-blocking")
chromeOptions.add_argument("--user-data-dir=" + profilePath)
#chromeOptions.add_argument("--profile-directory = Profile 1")
chromeOptions.add_experimental_option("prefs", { \
"profile.default_content_setting_values.media_stream_mic": 2,
"profile.default_content_setting_values.media_stream_camera": 2,
"profile.default_content_setting_values.geolocation": 2,
"profile.default_content_setting_values.notifications": 2
})

mailBoxXPath = '//*[@id="identifierId"]'
nextButtonXPath = '//*[@id="identifierNext"]/div/button'
enterPasswordBoxXPath = '//*[@id="password"]/div[1]/div/div[1]/input'
passwordNextButtonXPath = '//*[@id="passwordNext"]/div/button/div[2]'
meetLinkXPath = '//*[@id="yDmH0d"]/div[4]/div[3]/div/div[1]/div/div[2]/div[2]/div/span/a'
meetLinkInCommentsXPath = '//*[@id="ow43"]/div[2]/div/div[1]/div[2]/div[1]/html-blob/span/a[1]'
classroomPostClass = 'n8F6Jd'
meetLinkClass = 'qyN25' 
warningDismissButton = '//*[@id="yDmH0d"]/div[3]/div/div[2]/div[3]/div/span/span'
membersCountBeforeJoiningClass = 'Yi3Cfd'
joinButtonXPath = '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div/div[1]/div[1]/span/span'
captionsButtonXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[9]/div[3]/div[2]/div/span/span/div/div[1]/i'
captionsXPath = 'VbkSUe'
membersCountXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[1]/div[3]/div/div[2]/div[1]/span/span/div/div/span[2]'
chatBoxButtonXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[1]/div[3]/div/div[2]/div[3]/span/span'
chatBoxXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[4]/div/div[2]/div[2]/div[2]/span[2]/div/div[4]/div[1]/div[1]/div[2]/textarea'
chatSendButtonXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[4]/div/div[2]/div[2]/div[2]/span[2]/div/div[4]/div[2]/span/span'
chatBoxCloseXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[4]/div/div[2]/div[1]/div[2]/div/span/button/i'			


def toBold(str):
	boldedString = "\033[1m" + str + "\033[0m"
	return boldedString

def richStatus(text = 'Loading...', sleepTime = 10, spinnerType = 'dots', statusMessage = 'Done'):
	with console.status("[bold white] " + text, spinner = spinnerType) as status:
		time.sleep(sleepTime)
	console.print('[bold][green]' + statusMessage)

def printInSameLine(str1 = 'Loading', str2 = '.', sleepTime = 10, isChar = True, newLine = False, seconds = False):
	if newLine:
		print()
		return
	sec = sleepTime
	s = ' ' * 40
	for x in range (0, int(sleepTime) + 1): 
		b = str1 + (str(sec - x) if seconds else '') +  str(str2) * (x if isChar else 1)
		print(s, end = '\r')
		console.print(b, end = '\r', style = "bold green")
		if sleepTime > 0:
			time.sleep(1)
		s = ' ' * len(b) 
	if isChar:
		print('')

def findDay():
	dateAndTime = datetime.now()
	date = str(dateAndTime.day) + ' ' + str(dateAndTime.month) + ' ' + str(dateAndTime.year)
	date = datetime.strptime(date, '%d %m %Y').weekday()
	return date


def isSecondSaturday():
	dateAndTime = datetime.now()
	month = dateAndTime.month
	year = dateAndTime.year
	day = dateAndTime.day

	monthCalendar = calendar.monthcalendar(year, month)
	if monthCalendar[0][calendar.SATURDAY]:
		secondSaturday = monthCalendar[1][calendar.SATURDAY]
	else:
		secondSaturday = monthCalendar[2][calendar.SATURDAY]
	
	return secondSaturday


def updateholidaysList(key = None, value = None, remove = False):
	jsonData = fetchDataFromJSON('log.json')
	holidaysDict = jsonData["holidaysList"]
	if remove:
		del holidaysDict[key]
		console.print('Deleted '+ key + ' from holidays list successfully', style = "gold3")

	else :
		dateAndTime = datetime.now()
		day = dateAndTime.day
		secondSaturday = isSecondSaturday()
		if secondSaturday >= day:
			holidaysDict[str(secondSaturday)] = 'Second Saturday'
		if findDay() == 6:
			holidaysDict[str(day)] = 'Sunday'
		if not (key == None and value == None):
			holidaysDict[str(key)] = value
		for holidayDate in list(holidaysDict):
			if int(holidayDate) < day:
				holidaysDict.pop(holidayDate)

	jsonData["holidaysList"].update(holidaysDict)
	sendDataToJSON('log.json', jsonData)

def loadTimeTable():
	classTimeTableLocation = data['dir']['classTimeTableLocation']
	timetablewb = openpyxl.load_workbook(classTimeTableLocation) 
	sheet = timetablewb.active
	maxColumn = sheet.max_column
	maxRow = sheet.max_row
	completeTimeTable = {}
	for i in range(1, maxRow + 1):
		rowValues = []
		keyValue = sheet.cell(row = i, column = 1)
		for j in range(2, maxColumn + 1):
			cellValue = sheet.cell(row = i, column = j)
			if cellValue.value != None:
				rowValues.append(cellValue.value)
		completeTimeTable[keyValue.value] = rowValues

	jsonData = fetchDataFromJSON('log.json')
	jsonData["completeTimeTable"].update(completeTimeTable)
	sendDataToJSON('log.json', jsonData)

	classesToday()


def classesToday(printTable = False):
	date = findDay()
	classes = []
	updateholidaysList()
	jsonData = fetchDataFromJSON('log.json')
	holidays = jsonData["holidaysList"]
	dateAndTime = datetime.now()
	day = dateAndTime.day
	weekDay = dateAndTime.today().strftime('%A')

	if str(day) in holidays:
		console.print("\n\tToday is a holiday due to ", style = "bold cyan", end = '')
		console.print(holidays[str(day)] + '\n', style = "bold yellow")
	else:
		classes = []
		classtime = []
		classesToday = jsonData["completeTimeTable"][weekDay]
		timings = jsonData["completeTimeTable"]["Timings"]
		timings = timings[:len(classesToday)]

		prevClass = ''

		for i in range(len(classesToday)):
			if classesToday[i] == prevClass:
				classtime[-1] = classtime[-1][:8] + timings[i][8:]
			else :
				classes.append(classesToday[i])
				classtime.append(timings[i])
				prevClass = classesToday[i]

		t = {classtime[i]: classes[i] for i in range(len(classtime))}

		jsonData = fetchDataFromJSON('log.json')
		jsonData["todaysTimeTable"] = t
		sendDataToJSON('log.json', jsonData)

		if printTable:
			
			print(text2art("Classes today", font = "small"))
			currentTime = dateAndTime.time()
			time = str(currentTime).split(":")

			table = Table(show_lines=True)

			table.add_column("Timings", justify = "center", style = "cyan", no_wrap = True)
			table.add_column("Class", justify = "center", style = "green")
			table.add_column("Status", justify = "center", style = "magenta")
			

			classFlag = False
			for i in range(len(classtime)):
				if (time[0] >= classtime[i][0:2] and time[1] >= classtime[i][3:5]) and (time[0] < classtime[i][8:10] and time[1] <= '59'): 
					table.add_row(classtime[i], classes[i], "[bold magenta] ONGOING [green]:hourglass:")
					classFlag = True
				else:
					if classFlag == True:
						table.add_row(classtime[i], classes[i])
					else :
						table.add_row(classtime[i], classes[i], "[bold magenta]COMPLETED [green]:heavy_check_mark:")

			console.print(table)

def displayTimeTable():
	print(text2art("Time Table", font = "small"))
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	t = jsonData["completeTimeTable"]
	dateAndTime = datetime.now()
	weekDay = dateAndTime.today().strftime('%A')
	table = Table(show_lines=True)

	timetableKeys = list(t.keys())

	table.add_column(timetableKeys[0], justify = "center", style = "cyan", no_wrap = True)

	for i in range(len(t[timetableKeys[0]])):
		table.add_column(t[timetableKeys[0]][i], justify = "center", style = "yellow", no_wrap = True)

	for i in timetableKeys[1:]:
		tempList = [i] + t[i]
		if weekDay == i:
			table.add_row(*tempList, style = "bold")
		else:
			table.add_row(*tempList)
	
	console.print(table)

def displayHolidaysList():
	updateholidaysList()
	data = fetchDataFromJSON('log.json')
	holidayListKeys = list(data["holidaysList"].keys())
	table = Table(title = "[blink2 bold dodger_blue1]Holidays List", show_lines=True)
	table.add_column("[dark_orange]Date", justify = "center", style = "yellow3", no_wrap = True)
	table.add_column("[dark_orange]Occasion", justify = "center", style = "yellow3", no_wrap = True)
	holidaysFlag = False
	for key in holidayListKeys:
		value = data["holidaysList"][key]
		l = [key, value]
		table.add_row(*l)
		holidaysFlag = True

	if holidaysFlag:
		console.print(table)
	else :
		console.print("You dont have any holidays", style = "bold gold1")

def helpFunction():
	table = Table(show_lines=True)
	table.add_column("Arguments", justify = "center", style = "yellow3", no_wrap = True)
	table.add_column("Details", justify = "center", style = "yellow3")
	table.add_row("no arguments", "runs the main program")
	table.add_row("--t", "displays todays timetable")
	table.add_row("--h", "displays holidays list")
	table.add_row("--h -a", "add new holiday to the list and prints")
	table.add_row("--h -r", "removes holiday from the list and prints")
	table.add_row("--t -f", "displays complete timetable fetched from excel sheet")
	console.print(table)


def whichClass() :
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	subjects = jsonData["todaysTimeTable"]
	for i in subjects:
		time = datetime.now().time()
		time = str(time).split(":")
		if (time[0] >= i[0:2] and time[1] >= i[3:5]) and (time[0] < i[8:10] and time[1] <= '59'):
			return i[17:]		

	return None

def membersAlreadyJoinedCount(text):
	if text == 'No one else is here':
		return 0
	count = 0
	numberFromText = [int(i) for i in text.split() if i.isdigit()]
	commasCount = text.count(',')
	andCount = text.count('and')
	if len(numberFromText) > 0:
		count = numberFromText[0]
		andCount = 0

	count = count + commasCount + andCount + 1
	return count


def joinClass(subject, driver):
	subject = subject.upper()
	url = data['classroomLinks'][subject]
	print('Opening ' + subject + ' classroom in new tab' )
	driver.execute_script("window.open('');")
	driver.switch_to.window(driver.window_handles[1])
	driver.get(url)
	#printInSameLine(sleepTime = 5) #prints loading line by default
	richStatus(sleepTime = 5)
	print('Waiting for Google Meet link for ' + subject + ' class')

	usedPrintInSameLine = False
	linkPostedSeperatelyInAnnouncementTab = data['otherData']['linkPostedSeperatelyInAnnouncementTab']

	if subject in linkPostedSeperatelyInAnnouncementTab:	
		previousPostData = None
		while True:
			#From the below fetched data check the date is matching before joining
			announcementTabData = str(driver.find_element_by_class_name(classroomPostClass).text)
			if previousPostData != announcementTabData :
				print('Fetched Data \n')
				print(toBold(announcementTabData) + '\n')

			previousPostData = announcementTabData
			
			classURL = driver.find_element_by_xpath(meetLinkInCommentsXPath).text

			current_time = datetime.now() 

			# checking for date and month posted in announcementTabData
			if announcementTabData[9:11].lstrip('0') == str(current_time.day) and announcementTabData[12:14].lstrip('0') == str(current_time.month):
				if usedPrintInSameLine == True:
					printInSameLine(newLine = True)
				if (classURL[:24] == 'https://meet.google.com/') :
					print('Fetched ', classURL, 'from the google classroom')
					print('Opening ', classURL)
					driver.get(classURL)
					#printInSameLine(sleepTime = 5)
					richStatus(sleepTime = 5)
					break

				else:
					print('Fetching Link failed')
				
			else :
				driver.refresh()
				printInSameLine(str1 = 'Waiting for Todays link. Trying again in ', str2 = ' seconds', isChar = False, seconds = True)
				usedPrintInSameLine = True

	else :
		
		while True:
			try:
				if usedPrintInSameLine == True:
					printInSameLine(newLine = True)
				classData = driver.find_element_by_class_name(meetLinkClass).text
				classURL = re.search("(?P<url>https?://[^\s]+)", classData).group("url")
				print('classURL ', str(classURL))
				if classURL[:24] == 'https://meet.google.com/':
					print('Fetched ', classURL +'from the google classroom')
					print('Opening ', classURL)
					driver.get(classURL)
					richStatus(sleepTime = 5)
					print('Opened meet link')
					richStatus(sleepTime = 10)
					break
			except AttributeError:
				print('Meet link not available to fetch.')
				driver.refresh()
				printInSameLine(str1 = 'Trying again in ', str2 = ' seconds', isChar = False, seconds = True)
				usedPrintInSameLine = True


	print('Pressing dismiss button')
	warningDismiss = driver.find_element_by_xpath(warningDismissButton).click()
	time.sleep(3)

	membersCountBeforeJoiningData = driver.find_element_by_class_name(membersCountBeforeJoiningClass).text
	print('Members Joined\n')
	print(str(membersCountBeforeJoiningData), '\n')

	joinedMembers = membersAlreadyJoinedCount(membersCountBeforeJoiningData)

	usedPrintInSameLine = False
	minCountToJoin = data['otherData']['minCountToJoin']

	while True:
		if usedPrintInSameLine == True:
			printInSameLine(newLine = True)
		if joinedMembers >= minCountToJoin: 
			print('More than ' + str(minCountToJoin) + ' members already joined')
			print('Joining the class now')
			break
		else :
			if joinedMembers == 0:
				printInSameLine(str1 = 'No one joined. Trying again in ', str2 = ' seconds', isChar = False, seconds = True)
				usedPrintInSameLine = True
			else :
				print('Only ' + str(joinedMembers) + ' joined')
				print('Waiting for ' + str(minCountToJoin - joinedMembers) + ' more students to join the class')
				printInSameLine(str1 = 'Trying again in ', str2 = ' seconds', isChar = False, seconds = True)
				usedPrintInSameLine = True
				membersCountBeforeJoiningData = driver.find_element_by_class_name(membersCountBeforeJoiningClass).text
				joinedMembers = membersAlreadyJoinedCount(membersCountBeforeJoiningData)


	# clicks join button
	print('Pressing join button')
	join=driver.find_element_by_xpath(joinButtonXPath).click()
	time.sleep(3)


	# turn on captions
	print('Turning on captions')
	driver.find_element_by_xpath(captionsButtonXPath).click()
	time.sleep(4)

	# counting number of students joined 
	count = driver.find_element_by_xpath(membersCountXPath).text

	flag = False
	minCountToLeave = data['otherData']['minCountToLeave']
	alertWords = data['otherData']['alertWords']

	# Reads the text from captions until str(count) > '30':
	while True:
		count = driver.find_element_by_xpath(membersCountXPath).text
		printInSameLine('Members Count: ', count, sleepTime = 0, isChar = False)
		try:
			if count > str(minCountToLeave):
				flag = True
			elems = driver.find_element_by_class_name(captionsXPath)
			captionTextLower = str(elems.text).lower()

			for word in alertWords:
				if word in captionTextLower:
					printInSameLine(newLine = True)
					print(text2art("ALERT", font = "small")) 
					alertSound() # alert sound for soundCount times
					#responseMessage = data['otherData']['responseMessage']
					#sendMessageInChatBox(driver, responseMessage)
					
			if count < str(minCountToLeave) and flag :
				print('\nExiting Class')
				driver.close()
				break


		except (NoSuchElementException, StaleElementReferenceException):
			if count > str(minCountToLeave):
				flag = True
			if count < str(minCountToLeave) and flag :
				console.print('\nExiting Class', style = "blink2 bold red")
				driver.close()
				richStatus(sleepTime = 5, statusMessage = 'Left the class')
				break


def sendMessageInChatBox(driver, message):
	driver.find_element_by_xpath(chatBoxButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(1)
	chatBox = driver.find_element_by_xpath(chatBoxXPath)
	chatBox.send_keys(message)
	time.sleep(1)
	driver.find_element_by_xpath(chatSendButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(1)
	driver.find_element_by_xpath(chatBoxCloseXPath).click()
	driver.implicitly_wait(10)
	print('Responded to the class by sending ', toBold(responseMessage))
	richStatus(text = 'Message sent successfully', sleepTime = 10, spinnerType = 'point') 


def alertSound():
	beep = lambda x: os.system("echo -n '\a'; sleep 0.2;" * x)
	soundFrequency = data['otherData']['soundFrequency']
	beep(soundFrequency)
	print('Played alert sound successfully')
	richStatus(text = 'Played alert sound successfully', sleepTime = 10, spinnerType = 'point') 


def loadDriver():
	pathToChromeDriver = data['dir']['pathToChromeDriver']
	driver = webdriver.Chrome(options = chromeOptions, executable_path = pathToChromeDriver)
	driver.maximize_window()

	print('Disabled extensions')
	print('Turned off Location')
	console.print('Turned off Camera', style = "bold red")
	console.print('Turned off Microphone', style = "bold red")
	print('Turned off Pop-up')

	driver.get('https://classroom.google.com')
	print('Opening Google Classroom')
	richStatus(sleepTime = 5)

	return driver

	
def login():
	pathToChromeDriver = data['dir']['pathToChromeDriver']
	driver = webdriver.Chrome(options = chromeOptions, executable_path = pathToChromeDriver)
	driver.maximize_window()

	print('Disabled extensions')
	print('Turned off Location')
	console.print('Turned off Camera', style = "bold red")
	console.print('Turned off Microphone', style = "bold red")
	print('Turned off Pop-up')

	print('Logging into ' + toBold('Google account'))
	driver.get('https://classroom.google.com/?emr=0')
	time.sleep(3)

	mailAddress = data['credentials']['mailAddress']
	password = data['credentials']['password']

	print('Entering mail address')
	mailBox = driver.find_element_by_xpath(mailBoxXPath)
	driver.implicitly_wait(10)
	mailBox.send_keys(mailAddress)
	driver.find_element_by_xpath(nextButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(2)

	print(toBold('Entering password'))
	passwordBox = driver.find_element_by_xpath(enterPasswordBoxXPath)
	driver.implicitly_wait(10)
	passwordBox.send_keys(password)
	driver.find_element_by_xpath(passwordNextButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(2)

	printInSameLine(sleepTime = 10)
	print('Login Successful')

	return driver

		